#!/usr/bin/env python3
"""
Excel Utilities
Core module for local XLSX manipulation
"""

import os
import json
from pathlib import Path
from typing import Optional, List, Dict, Any, Union

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.utils import get_column_letter, column_index_from_string
except ImportError:
    print("openpyxl not installed. Run: pip install openpyxl")
    load_workbook = None


class XLSXAnalyzer:
    """Analyze Excel files"""

    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"XLSX not found: {file_path}")
        self.wb = load_workbook(str(self.file_path), data_only=False)

    def get_metadata(self) -> Dict[str, Any]:
        """Get workbook metadata"""
        props = self.wb.properties

        return {
            'file': str(self.file_path),
            'sheets': self.wb.sheetnames,
            'sheet_count': len(self.wb.sheetnames),
            'title': props.title or 'N/A',
            'creator': props.creator or 'N/A',
            'created': str(props.created) if props.created else 'N/A',
            'modified': str(props.modified) if props.modified else 'N/A'
        }

    def get_sheet_info(self, sheet_name: str) -> Dict[str, Any]:
        """Get info for specific sheet"""
        if sheet_name not in self.wb.sheetnames:
            raise ValueError(f"Sheet not found: {sheet_name}")

        ws = self.wb[sheet_name]

        # Get dimensions
        min_row = ws.min_row or 1
        max_row = ws.max_row or 1
        min_col = ws.min_column or 1
        max_col = ws.max_column or 1

        # Count cells with data
        data_cells = 0
        formula_cells = 0

        for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                                 min_col=min_col, max_col=max_col):
            for cell in row:
                if cell.value is not None:
                    data_cells += 1
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        formula_cells += 1

        return {
            'name': sheet_name,
            'rows': max_row,
            'columns': max_col,
            'range': f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}",
            'data_cells': data_cells,
            'formula_cells': formula_cells
        }

    def read_range(
        self,
        sheet_name: str,
        range_str: str
    ) -> List[List[Any]]:
        """Read data from a range"""
        ws = self.wb[sheet_name]

        data = []
        for row in ws[range_str]:
            row_data = []
            for cell in row:
                row_data.append(cell.value)
            data.append(row_data)

        return data

    def list_formulas(self, sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
        """List all formulas"""
        formulas = []

        sheets = [sheet_name] if sheet_name else self.wb.sheetnames

        for sname in sheets:
            ws = self.wb[sname]
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        formulas.append({
                            'sheet': sname,
                            'cell': cell.coordinate,
                            'formula': cell.value
                        })

        return formulas

    def analyze(self) -> Dict[str, Any]:
        """Full workbook analysis"""
        metadata = self.get_metadata()

        analysis = {
            **metadata,
            'sheets_detail': [],
            'total_formulas': 0
        }

        for sheet_name in self.wb.sheetnames:
            sheet_info = self.get_sheet_info(sheet_name)
            analysis['sheets_detail'].append(sheet_info)
            analysis['total_formulas'] += sheet_info['formula_cells']

        return analysis


class XLSXEditor:
    """Edit Excel files"""

    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"XLSX not found: {file_path}")
        self.wb = load_workbook(str(self.file_path))

    def update_cell(
        self,
        sheet_name: str,
        cell: str,
        value: Any,
        is_formula: bool = False
    ):
        """Update single cell"""
        ws = self.wb[sheet_name]

        if is_formula and not str(value).startswith('='):
            value = f"={value}"

        ws[cell] = value

    def update_range(
        self,
        sheet_name: str,
        start_cell: str,
        values: List[List[Any]]
    ):
        """Update a range of cells"""
        ws = self.wb[sheet_name]

        # Parse start cell
        col_letter = ''.join(filter(str.isalpha, start_cell))
        start_row = int(''.join(filter(str.isdigit, start_cell)))
        start_col = column_index_from_string(col_letter)

        for row_idx, row_values in enumerate(values):
            for col_idx, value in enumerate(row_values):
                cell = ws.cell(
                    row=start_row + row_idx,
                    column=start_col + col_idx
                )
                cell.value = value

    def find_replace(
        self,
        find: str,
        replace: str,
        sheet_name: Optional[str] = None
    ) -> int:
        """Find and replace text"""
        count = 0

        sheets = [self.wb[sheet_name]] if sheet_name else self.wb.worksheets

        for ws in sheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        if find in cell.value:
                            cell.value = cell.value.replace(find, replace)
                            count += 1

        return count

    def add_row(
        self,
        sheet_name: str,
        position: int,
        values: List[Any]
    ):
        """Insert row at position"""
        ws = self.wb[sheet_name]
        ws.insert_rows(position)

        for col_idx, value in enumerate(values, 1):
            ws.cell(row=position, column=col_idx, value=value)

    def delete_rows(
        self,
        sheet_name: str,
        start_row: int,
        count: int = 1
    ):
        """Delete rows"""
        ws = self.wb[sheet_name]
        ws.delete_rows(start_row, count)

    def add_sheet(self, name: str, position: Optional[int] = None):
        """Add new sheet"""
        if position is not None:
            return self.wb.create_sheet(name, position)
        return self.wb.create_sheet(name)

    def delete_sheet(self, name: str):
        """Delete sheet"""
        if name in self.wb.sheetnames:
            del self.wb[name]

    def copy_sheet(self, source_name: str, new_name: str):
        """Copy sheet"""
        source = self.wb[source_name]
        target = self.wb.copy_worksheet(source)
        target.title = new_name

    def save(self, output_path: str) -> str:
        """Save workbook"""
        output = Path(output_path)
        output.parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(str(output))
        return str(output)


def analyze_xlsx(file_path: str) -> Dict[str, Any]:
    """Quick analysis function"""
    analyzer = XLSXAnalyzer(file_path)
    return analyzer.analyze()


def read_range(file_path: str, sheet: str, range_str: str) -> List[List[Any]]:
    """Read data from range"""
    analyzer = XLSXAnalyzer(file_path)
    return analyzer.read_range(sheet, range_str)


def update_cell(
    file_path: str,
    sheet: str,
    cell: str,
    value: Any,
    output_path: str
) -> str:
    """Update single cell"""
    editor = XLSXEditor(file_path)
    editor.update_cell(sheet, cell, value)
    return editor.save(output_path)


def find_replace(
    file_path: str,
    find: str,
    replace: str,
    output_path: str
) -> int:
    """Find and replace in XLSX"""
    editor = XLSXEditor(file_path)
    count = editor.find_replace(find, replace)
    editor.save(output_path)
    return count


if __name__ == '__main__':
    import sys

    if len(sys.argv) < 2:
        print("Usage: python xlsx_utils.py <xlsx_file>")
        sys.exit(1)

    result = analyze_xlsx(sys.argv[1])
    print(json.dumps(result, indent=2, default=str))
